import csv
import json
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from sqlalchemy import case
from sqlalchemy.orm import Session

from ..api_utils import fail, ok
from ..database import get_db
from ..deps import require_manager_or_admin
from ..models import Activity, ActivityAssignment, Staff, User
from ..services.excel_service import sync_activity

router = APIRouter(prefix="/api/exports", tags=["exports"])

EXPORT_HEADERS = ["ID", "تاریخ", "نوع فعالیت", "نام مشتری", "آدرس", "شخص موظف", "وضعیت", "دستگاه", "گزارش", "سایر"]
IMPORT_HEADERS = ["ID", "تاریخ", "نوع فعالیت", "نام مشتری", "آدرس", "شخص موظف", "وضعیت", "دستگاه", "گزارش", "سایر"]


def _normalize_status(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    if text in {"done", "انجام شد", "انجام", "completed"}:
        return "done"
    return "pending"


def _normalize_address(address_value: str | None, location_value: str | None = None) -> str | None:
    address = (address_value or "").strip()
    if address:
        return address
    location = (location_value or "").strip()
    if location and location != "-":
        return location
    return None


def _normalize_location(address_value: str | None, location_value: str | None = None) -> str:
    address = _normalize_address(address_value, location_value)
    return address or "-"


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _parse_extra_fields(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    text = str(value).strip()
    if not text:
        return {}
    try:
        parsed = json.loads(text)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def _set_assignments(db: Session, activity: Activity, staff_ids: list[int], by_user_id: int) -> None:
    db.query(ActivityAssignment).filter(
        ActivityAssignment.activity_id == activity.id,
        ActivityAssignment.is_current.is_(True),
    ).update({"is_current": False})
    for sid in staff_ids:
        db.add(
            ActivityAssignment(
                activity_id=activity.id,
                staff_id=sid,
                assigned_by_user_id=by_user_id,
                is_current=True,
            )
        )


def _rows(db: Session):
    rows = (
        db.query(Activity)
        .order_by(case((Activity.status == "pending", 0), else_=1), Activity.priority.desc(), Activity.created_at.desc())
        .all()
    )
    result = []
    for a in rows:
        staff_names = []
        for x in db.query(ActivityAssignment).filter(ActivityAssignment.activity_id == a.id, ActivityAssignment.is_current.is_(True)):
            if x.staff:
                staff_names.append(x.staff.name)
        result.append(
            [
                a.id,
                a.date.isoformat(),
                a.activity_type,
                a.customer_name,
                a.address or "",
                ",".join(staff_names),
                "انجام شد" if a.status == "done" else "در انتظار",
                a.device_info or "",
                a.report_text or "",
                a.extra_fields_json or "{}",
            ]
        )
    return result


def _extract_staff_ids(raw_staff: str, staff_map: dict[str, int]) -> tuple[list[int], list[str]]:
    if not raw_staff:
        return [], []
    names = [x.strip() for x in raw_staff.replace("،", ",").split(",") if x.strip()]
    missing = [name for name in names if name.casefold() not in staff_map]
    ids = [staff_map[name.casefold()] for name in names if name.casefold() in staff_map]
    return ids, missing


def _parse_excel_rows(file_bytes: bytes, db: Session) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return [], [{"row": 1, "errors": ["فایل اکسل خالی است"]}]

    header_map: dict[str, int] = {}
    for idx, header in enumerate(header_row):
        key = str(header or "").strip()
        if key:
            header_map[key] = idx

    required_headers = {"تاریخ", "نوع فعالیت", "نام مشتری", "آدرس"}
    missing_headers = [x for x in required_headers if x not in header_map]
    if missing_headers:
        return [], [{"row": 1, "errors": [f"ستون های ضروری موجود نیست: {', '.join(missing_headers)}"]}]

    staff_rows = db.query(Staff).filter(Staff.active.is_(True)).all()
    staff_map = {s.name.casefold(): s.id for s in staff_rows}

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for row_idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = list(values)
        if not any(x is not None and str(x).strip() for x in raw):
            continue

        current_errors: list[str] = []

        raw_date = raw[header_map["تاریخ"]] if header_map.get("تاریخ") is not None else None
        raw_type = raw[header_map["نوع فعالیت"]] if header_map.get("نوع فعالیت") is not None else None
        raw_customer = raw[header_map["نام مشتری"]] if header_map.get("نام مشتری") is not None else None
        raw_address = raw[header_map["آدرس"]] if header_map.get("آدرس") is not None else None

        activity_date = _parse_date(raw_date)
        activity_type = str(raw_type or "").strip()
        customer_name = str(raw_customer or "").strip()
        address = _normalize_address(str(raw_address or "").strip())

        if not activity_date:
            current_errors.append("تاریخ معتبر نیست (فرمت YYYY-MM-DD)")
        if len(activity_type) < 2:
            current_errors.append("نوع فعالیت ضروری است")
        if len(customer_name) < 2:
            current_errors.append("نام مشتری ضروری است")
        if not address:
            current_errors.append("آدرس ضروری است")

        raw_staff = ""
        if "شخص موظف" in header_map:
            raw_staff = str(raw[header_map["شخص موظف"]] or "").strip()
        staff_ids, missing_staff = _extract_staff_ids(raw_staff, staff_map)
        if missing_staff:
            current_errors.append(f"کارمند یافت نشد: {', '.join(missing_staff)}")

        row_id: int | None = None
        if "ID" in header_map:
            raw_id = raw[header_map["ID"]]
            if raw_id is not None and str(raw_id).strip() != "":
                try:
                    row_id = int(str(raw_id).strip())
                except ValueError:
                    current_errors.append("ID باید عدد صحیح باشد")

        status = _normalize_status(raw[header_map["وضعیت"]] if "وضعیت" in header_map else None)
        device_info = str(raw[header_map["دستگاه"]] or "").strip() if "دستگاه" in header_map else ""
        report_text = str(raw[header_map["گزارش"]] or "").strip() if "گزارش" in header_map else ""
        extra_fields = _parse_extra_fields(raw[header_map["سایر"]] if "سایر" in header_map else None)

        parsed = {
            "row": row_idx,
            "id": row_id,
            "date": activity_date,
            "activity_type": activity_type,
            "customer_name": customer_name,
            "address": address,
            "location": _normalize_location(address),
            "status": status,
            "device_info": device_info or None,
            "report_text": report_text or None,
            "extra_fields": extra_fields,
            "staff_ids": staff_ids,
        }

        if current_errors:
            errors.append({"row": row_idx, "errors": current_errors})
        else:
            rows.append(parsed)

    return rows, errors


@router.get("/csv")
def export_csv(db: Session = Depends(get_db), user: User = Depends(require_manager_or_admin)):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_HEADERS)
    for row in _rows(db):
        writer.writerow(row)
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="activities.csv"'},
    )


@router.get("/excel")
def export_excel(db: Session = Depends(get_db), user: User = Depends(require_manager_or_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "فعالیت ها"
    ws.append(EXPORT_HEADERS)
    for row in _rows(db):
        ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="activities-export.xlsx"'},
    )


@router.get("/excel/template")
def export_excel_template(user: User = Depends(require_manager_or_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(IMPORT_HEADERS)
    ws.append(["", date.today().isoformat(), "نصب", "مشتری نمونه", "کابل", "", "pending", "", "", "{}"])

    bio = BytesIO()
    wb.save(bio)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="activities-import-template.xlsx"'},
    )


@router.post("/excel/validate")
async def validate_excel_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_manager_or_admin),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise fail("BAD_REQUEST", "فایل باید با پسوند .xlsx باشد", status_code=400)

    rows, errors = _parse_excel_rows(await file.read(), db)
    preview = [{"row": x["row"], "customer_name": x["customer_name"], "address": x["address"]} for x in rows[:50]]
    return ok(
        {
            "valid": len(errors) == 0,
            "total_rows": len(rows) + len(errors),
            "valid_rows": len(rows),
            "error_rows": len(errors),
            "errors": errors,
            "preview": preview,
        }
    )


@router.post("/excel/import")
async def import_excel(
    mode: str = Query(default="upsert", pattern="^(insert|upsert)$"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_manager_or_admin),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise fail("BAD_REQUEST", "فایل باید با پسوند .xlsx باشد", status_code=400)

    rows, errors = _parse_excel_rows(await file.read(), db)
    if errors:
        raise fail(
            "VALIDATION_ERROR",
            "فایل اکسل معتبر نیست",
            details={"error_rows": len(errors), "errors": errors},
            status_code=400,
        )

    created = 0
    updated = 0
    touched_ids: list[int] = []

    try:
        for item in rows:
            existing = None
            if mode == "upsert" and item["id"] is not None:
                existing = db.query(Activity).filter(Activity.id == item["id"]).first()

            if existing:
                existing.date = item["date"]
                existing.activity_type = item["activity_type"]
                existing.customer_name = item["customer_name"]
                existing.address = item["address"]
                existing.location = item["location"]
                existing.status = item["status"]
                existing.device_info = item["device_info"]
                existing.report_text = item["report_text"]
                existing.extra_fields_json = json.dumps(item["extra_fields"], ensure_ascii=False)
                if existing.status == "done" and not existing.done_at:
                    existing.done_at = datetime.utcnow()
                    existing.done_by_user_id = user.id
                if existing.status != "done":
                    existing.done_at = None
                    existing.done_by_user_id = None
                _set_assignments(db, existing, item["staff_ids"], user.id)
                updated += 1
                touched_ids.append(existing.id)
                continue

            row = Activity(
                id=item["id"] if mode == "upsert" and item["id"] is not None else None,
                created_by_user_id=user.id,
                date=item["date"],
                activity_type=item["activity_type"],
                customer_name=item["customer_name"],
                address=item["address"],
                location=item["location"],
                status=item["status"],
                priority=0,
                report_text=item["report_text"],
                device_info=item["device_info"],
                extra_fields_json=json.dumps(item["extra_fields"], ensure_ascii=False),
                done_by_user_id=user.id if item["status"] == "done" else None,
                done_at=datetime.utcnow() if item["status"] == "done" else None,
            )
            db.add(row)
            db.flush()
            _set_assignments(db, row, item["staff_ids"], user.id)
            created += 1
            touched_ids.append(row.id)

        db.commit()
    except Exception as exc:
        db.rollback()
        raise fail("IMPORT_FAILED", "وارد کردن اکسل ناموفق بود", details=str(exc), status_code=500) from exc

    for activity_id in touched_ids:
        sync_activity(db, activity_id)

    return ok({"mode": mode, "created": created, "updated": updated, "imported": len(rows), "activity_ids": touched_ids})
