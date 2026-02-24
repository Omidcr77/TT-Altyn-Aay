import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from ..config import settings
from ..models import Activity

HEADERS = [
    "ID",
    "تاریخ",
    "نوع فعالیت",
    "نام مشتری",
    "آدرس",
    "شخص موظف",
    "وضعیت",
    "دستگاه / امانت",
    "گزارش",
    "سایر معلومات",
    "ایجاد شده",
    "تکمیل شده",
]


def _get_sheet(path: Path) -> tuple[Workbook, Worksheet]:
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "فعالیت ها"
    ws.append(HEADERS)
    wb.save(path)
    return wb, ws


def ensure_excel_exists() -> None:
    _get_sheet(settings.excel_file)[0].save(settings.excel_file)


def _activity_row(activity: Activity) -> list[str]:
    assigned = ", ".join(a.staff.name for a in activity.assignments if a.is_current and a.staff is not None)
    return [
        str(activity.id),
        activity.date.isoformat(),
        activity.activity_type,
        activity.customer_name,
        activity.address or "",
        assigned,
        "انجام شد" if activity.status == "done" else "در انتظار",
        activity.device_info or "",
        activity.report_text or "",
        activity.extra_fields_json or "{}",
        activity.created_at.strftime("%Y-%m-%d %H:%M"),
        activity.done_at.strftime("%Y-%m-%d %H:%M") if activity.done_at else "",
    ]


def sync_activity(db: Session, activity_id: int) -> None:
    activity = db.query(Activity).filter(Activity.id == activity_id).first()
    if not activity:
        return
    _ = json.loads(activity.extra_fields_json or "{}")
    path = settings.excel_file
    wb, ws = _get_sheet(path)
    target_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == str(activity.id):
            target_row = row
            break
    values = _activity_row(activity)
    if target_row is None:
        ws.append(values)
    else:
        for col, value in enumerate(values, start=1):
            ws.cell(row=target_row, column=col, value=value)
    wb.save(path)
